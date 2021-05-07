from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from tkinter import *
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
import mmap
from random import randint
import csv
import tkinter as tk
from tkinter import filedialog
import tkinter.font as font
import os
import ntpath
import sys


class ScraperUI():
    def __init__(self):
        self.root = tk.Tk()
        self.root.geometry("300x250")
        self.root.title("Select Your Folder To Check Whether The Titles are Listed in Amazon")
        self.uiFont = font.Font(size=20)
        self.google_group = None
        self.google_group_box = None
        self.browser = None
        self.SetBrowserLabel()
        self.SetBrowserPreference()
        self.SetSaveBrowserButton()
        self.SetGoogleGroupLabel()
        self.SetGoogleGroupPreference()
        self.SetSaveGoogleGroupButton()
        self.root.after(0, self.checkValues)
        self.root.mainloop()

    def checkValues(self, debug=False):
        if debug:
            print(self.google_group, self.browser)
        if self.google_group and self.browser:
            time.sleep(1)
            self.root.destroy()
        self.root.after(100, self.checkValues)

    def SetGoogleGroupLabel(self):
        self.google_group_label = tk.Label(self.root, text="Select The Google Group you want to scrape ")
        self.google_group_label.pack()

    def SetGoogleGroup(self):
        self.google_group = str(self.googleGroupVal.get())

    def SetGoogleGroupPreference(self):
        self.googleGroupVal = StringVar()
        self.vitians2021 = tk.Radiobutton(self.root, text="vitians2021", variable=self.googleGroupVal,
                                          value='vitians02021',
                                          command=self.SetGoogleGroup)
        self.vitians2021.select()
        self.vitians2021.pack()
        self.vitians2020 = tk.Radiobutton(self.root, text="vitians2020", variable=self.googleGroupVal,
                                          value='vitians2020',
                                          command=self.SetGoogleGroup)
        self.vitians2020.deselect()
        self.vitians2020.pack()

    def SetSaveGoogleGroupButton(self):
        self.saveGoogleGroupButton = tk.Button(self.root, bg='yellow', text='Save GoogleGroup',
                                               command=self.SetGoogleGroup)
        self.saveGoogleGroupButton.configure(width=600, height=2)
        self.saveGoogleGroupButton.pack()

    def SetBrowserLabel(self):
        self.browser_label = tk.Label(self.root, text="Select The Your Preferred Browser ")
        self.browser_label.pack()

    def SetBrowserPreference(self):
        self.BrowserVal = StringVar()
        self.firefoxRadioButton = tk.Radiobutton(self.root, text="Firefox", variable=self.BrowserVal, value='Firefox',
                                                 command=self.SetBrowser)
        self.firefoxRadioButton.deselect()
        self.firefoxRadioButton.pack()
        self.chromeRadioButton = tk.Radiobutton(self.root, text="Chrome", variable=self.BrowserVal, value='Chrome',
                                                command=self.SetBrowser)
        self.chromeRadioButton.select()
        self.chromeRadioButton.pack()

    def SetSaveBrowserButton(self):
        self.saveBrowserButton = tk.Button(self.root, bg='yellow', text='Save Browser', command=self.SetBrowser)
        self.saveBrowserButton.configure(width=600, height=2)
        self.saveBrowserButton.pack()

    def SetBrowser(self):
        self.browser = str(self.BrowserVal.get())

    def GetGoogleGroup(self):
        return self.google_group

    def GetBrowserPreference(self):
        return self.browser


class PlacementScraper():

    def __init__(self,debug=False):
        self.scraperUI = ScraperUI()
        self.browser = self.scraperUI.GetBrowserPreference()
        self.google_group = self.scraperUI.GetGoogleGroup()
        if self.google_group:
            if self.browser == 'Firefox':
                self.SetupFirefoxDriver()
            elif self.browser == 'Chrome':
                self.SetupChromeDriver()
            self.first_page = True
            self.headers_row = ['Name of the Company', 'Category', 'Date of Visit: ', 'CTC', 'Stipend']
            self.all_registration_links = []
            self.filename = 'placementInfo' + self.google_group + '.xlsx'
            self.GetYearFromGoogleGroup()
            self.CreateRegistrationSearchText()
            self.CreateOutputFile()
            if debug:
                self.test()
            else:
                self.next_page_element_class = 'uArJ5e Y5FYJe cjq2Db YLAhbd M9Bg4d'
                self.OpenVitiansGroupLink()
                while not self.CheckNextButtonDisabled() or self.first_page:
                    if self.first_page:
                        self.first_page = False
                    self.GetAllRegistrationLinks()
                    self.GetNextPage()

            self.DestroyDriver()

    def CreateRegistrationSearchText(self):
        self.registration_search_texts = ['placement Registration ' + str(self.year) + ' Batch',
                                          'Registration Notice -']

    def CheckNextButtonDisabled(self):
        try:
            self.next_page_element = self.driver.find_element_by_css_selector('div[aria-label="Next page"]')
            self.next_page_exists = self.next_page_element.get_attribute('aria-disabled')
            if self.next_page_exists == 'false':
                return False
            else:
                return True
        except NoSuchElementException:
            return True

    def CreateOutputFile(self):
        if not os.path.exists(self.filename):
            output_workbook = Workbook()
            output_worksheet = output_workbook.active
            output_worksheet.append(self.headers_row)
            output_workbook.save(self.filename)

    def OpenVitiansGroupLink(self):
        self.link = 'https://groups.google.com/g/' + self.google_group
        self.driver.get(self.link)

    def GetNextPage(self):
        self.next_page_element = self.driver.find_element_by_css_selector('div[aria-label="Next page"]')
        self.next_page_element.click()
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div[class="NUZAVc L6cTce"]')))

    def GetYearFromGoogleGroup(self):
        self.year = ''
        for digit in self.google_group:
            if digit.isdigit():
                self.year += str(digit)
        if len(self.year) == 5:
            self.year = self.year[1:]

    def GetAllRegistrationLinks(self):
        self.all_registration_elements = self.driver.find_elements_by_css_selector(
            'span[class=" eois5  ETPPOd USV3fe CzRzXb dwBwye"')
        for webpage in self.all_registration_elements:
            for self.registration_search_text in self.registration_search_texts:
                if self.registration_search_text in webpage.text:
                    webpage.click()
                    overlay_element = self.driver.find_element_by_css_selector(
                        'c-wiz[class="zQTmif SSPGKf eejsDc oCHqfe"]')
                    style_attr = overlay_element.get_attribute('style')
                    while 'visibility: hidden; opacity: 0;' in str(style_attr):
                        overlay_element = self.driver.find_element_by_css_selector(
                            'c-wiz[class="zQTmif SSPGKf eejsDc oCHqfe"]')
                        style_attr = overlay_element.get_attribute('style')
                    time.sleep(2)
                    self.GetTableContents()
                    self.WriteRecord()
                    self.PrintRecord()
                    self.driver.back()

    def GetTableContents(self):
        try:
            self.main_body = self.driver.find_element_by_class_name('ptW7te')
            self.table_element = self.main_body.find_element_by_tag_name('table')
        except NoSuchElementException:
            self.main_body = self.driver.find_element_by_class_name('gmail_quote')
            self.table_element = self.main_body.find_element_by_tag_name('table')
        self.table_rows = self.table_element.find_elements_by_tag_name('tr')
        self.row_information = ['', '', '', '', '']
        for row in self.table_rows:
            for header_num in range(len(self.headers_row)):
                if self.headers_row[header_num] in str(row.text):
                    self.row_information[header_num] = str(row.text).replace(self.headers_row[header_num], '').strip()
                    break
        if self.row_information[3] == '':
            for row in self.table_rows:
                if '13.9' in str(row.get_attribute('style')) :
                    self.row_information[3] += str(row.text)
        self.row_information[3] = self.row_information[3].replace('CTC', '').strip()

    def test(self):
        self.test_link = 'https://groups.google.com/g/vitians02021/c/aA_-lxRig8s/m/ir5RruNdAQAJ'
        self.driver.get(self.test_link)
        self.GetTableContents()
        self.WriteRecord()
        self.PrintRecord()
        self.driver.back()

    def SetupChromeDriver(self):
        self.options = Options()
        self.options.add_argument('--no-sandbox')
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.chrome_driver_path = r"chromedriver.exe"
        self.driver = webdriver.Chrome(executable_path=self.chrome_driver_path, options=self.options)

    def SetupFirefoxDriver(self):
        self.options = FirefoxOptions()
        self.options.add_argument('--no-sandbox')
        self.firefox_driver_path = r"geckodriver.exe"
        self.driver = webdriver.Firefox(executable_path=self.firefox_driver_path, options=self.options)

    def DestroyDriver(self):
        try:
            self.driver.quit()
            self.driver.close()
        except:
            pass

    def WriteRecord(self):
        if os.path.exists(self.filename):
            output_workbook = load_workbook(self.filename)
            output_worksheet = output_workbook.worksheets[0]
            output_worksheet.append(self.row_information)
            output_workbook.save(self.filename)

    def PrintRecord(self):
        for num in range(len(self.headers_row)):
            print(self.headers_row[num], self.row_information[num], sep=" : ", end=' , ')
        print()


Ps = PlacementScraper()
